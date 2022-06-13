from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class ExcelManipulate:
    def __init__(self, path:str):
        self.path = path
        self.wb = load_workbook(path)
        self.ws = self.wb.active

    # sheet 預設值為1 表示第一個sheet
    def getCellcontent(self, row:int, column:int, sheet=1):
        worksheet = self.wb.worksheets[sheet-1]
        targetIndex = get_column_letter(column) + str(row)
        return str(worksheet[targetIndex].value)

    def getRowDataList(self, row:int):
        resultList =[]
        for i in range(1, self.ws.max_column+1):
            eachCell = self.getCellcontent(row, i)
            resultList.append(eachCell)

        return resultList

    def getAllDataList(self):
        resultList =[]
        for i in range(1, self.ws.max_row+1):
            eachRow = self.getRowDataList(i)
            resultList.append(eachRow)

        return resultList

    # 全列搜尋
    def compareContentByColumnAllRow(self,compareContent:str, column:int, sheetIndex=1): # 比較某一欄位資料，傳回符合的值的 row Index 的 list
        resultList=[]
        targetSheet = self.wb.worksheets[sheetIndex-1]

        for i in range(1, targetSheet.max_row+1):
            if str(self.getCellcontent(i, column)) == compareContent:
                resultList.append(i)

        return resultList

    def compareContentByColumnWithinRow(self,compareContent:str, column:int, rowList:list): # 比較某一欄位資料，傳回符合的值的 row Index 的 list
        resultList=[]
        if len(rowList) == 0 : return resultList
        for i in rowList:
            if str(self.getCellcontent(i, column)) == compareContent:
                resultList.append(i)

        return resultList


    def writeBackToExcel(self, row:int, column:int ,resultString:str):
        self.ws[get_column_letter(column)+str(row)].value = resultString
        self.wb.save(self.path)